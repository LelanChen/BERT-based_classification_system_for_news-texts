# coding: UTF-8

import time
import openpyxl
import os
import torch
import tqdm
import numpy as np
from openpyxl import Workbook

from need.utils import get_time_dif
from need.predicate_utils import load_dataset_one, build_iterator_one,load_dataset_nums,load_dataset_excel, build_iterator_nums,load_dataset_csv


def _predicate(config,model,predict_iter):
    model.load_state_dict(torch.load(config.save_path))  # 加载保存的模型
    model.eval()
    predict_all = np.array([], dtype=int)
    with torch.no_grad():
        for texts in predict_iter:
            outputs = model(texts)
            predict = torch.max(outputs.data, 1)[1].cpu().numpy()
            predict_all = np.append(predict_all, predict)

    num = len(predict_all)  # 预测结果数目
    print("分类结果：")
    result_all=[]
    for i in range(0, num):
        lab = int(predict_all[i])
        class_result = config.class_list[lab]
        print(class_result)  # 打印出所有的预测结果
        result_all.append(class_result)

    return result_all

def save_histry(new,result,save_dir='Histry/histry.xlsx'):
    '''
    保存历史记录（若输入的是表单直接将结果保存至原表单，不另外保存至历史记录）
    new:预测的新闻内容
    result：预测分类结果
    save_dir:历史记录保存的地址，默认为'Histry/histry.xlsx'
    '''
    def save(new,result,wb):
        for i in range(0,len(new)):
            sheet_name = wb.sheetnames  # 将excel所有表单名称读出
            if result[i] in sheet_name:
                ws = wb[result[i]]
                ws.cell(ws.max_row + 1, 1).value = new[i]
                ws.cell(ws.max_row , 2).value = result[i]
            else:
                ws1 = wb.create_sheet(title=result[i])
                ws1.cell(1, 1).value = 'content'
                ws1.cell(1, 2).value = 'channelName'
                ws1.cell(2, 1).value = new[i]
                ws1.cell(2, 2).value = result[i]
    try:
        wb = openpyxl.load_workbook(save_dir)  # 打开历史记录文件
        save(new, result, wb)
    except FileNotFoundError:
        #创建工作簿
        wb = Workbook()
        save(new, result, wb)
    wb.save(save_dir)


def predicate_one(config, model, text=None):
    '''
    predicate_one()调用训练好的模型预测单条新闻
    text:一条新闻内容
    histry:预测后新闻的历史记录存储地址
    '''
    start_time = time.time()
    print("Loading data...")
    predict_data ,new= load_dataset_one(config, text, config.pad_size)
    predict_iter = build_iterator_one(predict_data, config)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    class_result=_predicate(config,model,predict_iter)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    #保存分类结果至历史记录
    save_histry(new, class_result)
    return class_result[0]

def predicate_nums(config, model, text_dir,histry='Histry/histry' ):
    '''
    predicate_nums()调用训练好的模型预测多条新闻
    text_dir:存放多条新闻文本的一个文件夹地址
    histry:预测后新闻的历史记录存储地址
    '''
    start_time = time.time()
    print("Loading data...")
    predict_data_nums,news = load_dataset_nums(config, text_dir, config.pad_size)
    predict_iter_nums = build_iterator_nums(predict_data_nums, config)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    result_all = _predicate(config,model, predict_iter_nums)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    #保存分类结果至历史记录
    save_histry(news, result_all)
    #将预测新闻和分类结果保存至原文件夹内供用户查看
    save_dir=text_dir + '/' + 'result.xlsx'
    save_histry(news, result_all,save_dir)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)

def predicate_excel(config, model, excel_dir ):
    '''
    predicate_nums()调用训练好的模型预测多条新闻
    excel_dir:存放多条新闻文本的一个excel地址
    '''
    start_time = time.time()
    print("Loading data...")
    wb = openpyxl.load_workbook(excel_dir)  # 打开excel文件
    ws = wb.active
    predict_data_excel = load_dataset_excel(config, ws, config.pad_size)
    print(len(predict_data_excel))
    predict_iter_excel = build_iterator_nums(predict_data_excel, config)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    result = _predicate(config, model, predict_iter_excel)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    # 保存分类结果至表单中
    if ws.cell(2, 2).value is None:
        for i in range(2, ws.max_row + 1):
            ws.cell(i, 2).value = result[i-2]
    else:
        for i in range(2, ws.max_row + 1):
            if i == 2:
                c = ws.max_column+1
            ws.cell(i, c).value = result[i-2]
    wb.save(excel_dir)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)

def predicate_csv(config, model, csv_dir ):
    '''
    predicate_nums()调用训练好的模型预测多条新闻
    csv_dir:存放多条新闻文本的一个csv文件地址
    '''
    start_time = time.time()
    print("Loading data...")
    predict_data_csv = load_dataset_csv(config, csv_dir, config.pad_size)
    print(len(predict_data_csv))
    predict_iter_csv = build_iterator_nums(predict_data_csv, config)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    result = _predicate(config, model, predict_iter_csv)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)
    # 保存分类结果至表单中
    with open(csv_dir, 'r') as csvFile:
        rows = csv.reader(csvFile)
        with open("Histry/result.csv", 'w', newline='') as f:
            writer = csv.writer(f)
            i = 1
            n = 0
            print(len(result))
            for row in rows:
                if i == 1:
                    row.append('class')
                    i = i + 1
                else:
                    row.append(result[n])
                    print(n)
                    n = n + 1
                writer.writerow(row)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)


if __name__ == '__main__':
    '''
    调试该功能模块时使用
    '''
    #model_name = args.model  # bert
    #x = import_module('models.' + model_name)
    from models import preedicate_config as x
    config = x.Config()
    np.random.seed(1)
    torch.manual_seed(1)
    torch.cuda.manual_seed_all(1)
    torch.backends.cudnn.deterministic = True  # 保证每次结果一样

    # predict
    text = '上海证券报＊ＳＴ中钨股票近日交易异常波动。２００８年１月７日，公司股票因重大无先例事项实施停牌，至２００８年２月１８日，经向控股股东湖南有色金属股份有限公司核实后，由于相关方案尚不成熟且无实质性进展，公司股票予以复牌。'
    text_dir='THUCNews/pre_news'
    excel_dir='THUCNews/data/test_4.xlsx'
    model = x.Model(config).to(config.device)
    #predicate_one(config, model, text)
    #predicate_nums(config,model,text_dir)
    #predicate_excel(config, model, excel_dir)